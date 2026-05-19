# 资产管理路由：资产 CRUD、二维码生成、Excel 导入导出、时间线、保修提醒、折旧计算
# 资产编号自动生成规则：IT-YYYYMM-NNNN（年、月、4位流水号）

import io
import csv
import random
import string
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from database import get_db
from auth import get_current_user, require_permission, require_dept_scope
from schemas import AssetCreate, AssetUpdate, Response
import qrcode
import openpyxl

router = APIRouter(prefix="/api/assets", tags=["资产管理"])


def _format_asset(row):
    return {
        "id": row["id"], "asset_no": row["asset_no"], "name": row["name"],
        "category_id": row["category_id"], "brand": row["brand"], "model": row["model"],
        "serial_no": row["serial_no"], "purchase_price": row["purchase_price"],
        "purchase_date": row["purchase_date"], "status": row["status"],
        "dept_id": row["dept_id"], "user_id": row["user_id"],
        "warehouse_id": row["warehouse_id"], "location": row["location"],
        "supplier_id": row["supplier_id"], "warranty_date": row["warranty_date"],
        "remark": row["remark"], "create_time": row["create_time"], "update_time": row["update_time"],
        "purchase_lifespan_years": row["purchase_lifespan_years"] if "purchase_lifespan_years" in row.keys() else 0,
        "depreciation_method": row["depreciation_method"] if "depreciation_method" in row.keys() else "straight",
        "monthly_depreciation": row["monthly_depreciation"] if "monthly_depreciation" in row.keys() else 0,
        "accumulated_depreciation": row["accumulated_depreciation"] if "accumulated_depreciation" in row.keys() else 0,
        "net_value": row["net_value"] if "net_value" in row.keys() else 0,
    }


def _gen_asset_no(db) -> str:
    now = datetime.now()
    ym = now.strftime("%Y%m")
    row = db.execute(
        "SELECT asset_no FROM assets WHERE asset_no LIKE ? ORDER BY id DESC LIMIT 1",
        (f"IT-{ym}-%",)
    ).fetchone()
    if row:
        seq = int(row["asset_no"].split("-")[-1]) + 1
    else:
        seq = 1
    return f"IT-{ym}-{seq:04d}"


def _log(db, user_id: int, desc: str):
    db.execute("INSERT INTO operation_logs (user_id, description) VALUES (?, ?)", (user_id, desc))
    db.commit()


@router.get("")
def list_assets(
    keyword: str = Query(None), category_id: int = Query(None),
    status: str = Query(None), dept_id: int = Query(None), warehouse_id: int = Query(None),
    page: int = Query(1), page_size: int = Query(20),
    format: str = Query(None),
    user: dict = Depends(require_permission("asset:read")),
    scope: dict = Depends(require_dept_scope())
):
    db = get_db()
    conditions = []
    params = []
    if keyword:
        conditions.append("(a.name LIKE ? OR a.asset_no LIKE ? OR a.serial_no LIKE ?)")
        params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
    if category_id:
        conditions.append("a.category_id = ?")
        params.append(category_id)
    if status:
        conditions.append("a.status = ?")
        params.append(status)
    if dept_id:
        conditions.append("a.dept_id = ?")
        params.append(dept_id)
    if warehouse_id:
        conditions.append("a.warehouse_id = ?")
        params.append(warehouse_id)
    if scope:
        for k, v in scope.items():
            conditions.append(f"a.{k} = ?")
            params.append(v)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    # CSV 导出
    if format == "csv":
        rows = db.execute(
            f"""SELECT a.*, c.name as category_name, d.name as dept_name, w.name as warehouse_name
               FROM assets a
               LEFT JOIN categories c ON a.category_id = c.id
               LEFT JOIN departments d ON a.dept_id = d.id
               LEFT JOIN warehouses w ON a.warehouse_id = w.id
               {where} ORDER BY a.id DESC""",
            params
        ).fetchall()
        db.close()
        output = io.StringIO()
        writer = csv.writer(output)
        status_map = {"in_stock": "在库", "in_use": "使用中", "borrowed": "借出", "repairing": "维修中", "scrapped": "已报废"}
        writer.writerow(["资产编号", "名称", "分类", "品牌", "型号", "序列号", "采购价格", "采购日期",
                          "状态", "部门", "仓库", "存放位置", "保修到期", "备注"])
        for r in rows:
            writer.writerow([r["asset_no"], r["name"], r["category_name"], r["brand"], r["model"],
                             r["serial_no"], r["purchase_price"], r["purchase_date"],
                             status_map.get(r["status"], r["status"]),
                             r["dept_name"], r["warehouse_name"], r["location"], r["warranty_date"], r["remark"]])
        output.seek(0)
        suffix = ''.join(random.choices(string.ascii_lowercase + string.digits, k=5))
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv; charset=utf-8",
                                 headers={"Content-Disposition": f"attachment; filename=assets_{suffix}.csv"})

    total = db.execute(f"SELECT COUNT(*) FROM assets a {where}", params).fetchone()[0]
    offset = (page - 1) * page_size
    rows = db.execute(
        f"""SELECT a.*, c.name as category_name, d.name as dept_name, w.name as warehouse_name,
                  (SELECT COUNT(*) FROM repairs r WHERE r.asset_id = a.id) as repair_count
           FROM assets a
           LEFT JOIN categories c ON a.category_id = c.id
           LEFT JOIN departments d ON a.dept_id = d.id
           LEFT JOIN warehouses w ON a.warehouse_id = w.id
           {where} ORDER BY a.id DESC LIMIT ? OFFSET ?""",
        params + [page_size, offset]
    ).fetchall()
    items = []
    for r in rows:
        item = _format_asset(r)
        item["category_name"] = r["category_name"]
        item["dept_name"] = r["dept_name"]
        item["warehouse_name"] = r["warehouse_name"]
        item["repair_count"] = r["repair_count"]
        items.append(item)
    db.close()
    return Response(data={"total": total, "page": page, "page_size": page_size, "items": items}).model_dump()


@router.get("/names")
def get_asset_names(user: dict = Depends(get_current_user)):
    db = get_db()
    rows = db.execute("SELECT DISTINCT name FROM assets ORDER BY name").fetchall()
    db.close()
    return Response(data=[r["name"] for r in rows]).model_dump()


@router.get("/warranty-alerts")
def warranty_alerts(days: int = Query(30), user: dict = Depends(require_permission("asset:read"))):
    """返回保修到期日在 [today, today+days] 范围内的资产列表"""
    db = get_db()
    today = date.today()
    end_date = (today + timedelta(days=days)).isoformat()
    rows = db.execute(
        """SELECT a.*, c.name as category_name
           FROM assets a LEFT JOIN categories c ON a.category_id = c.id
           WHERE a.warranty_date >= ? AND a.warranty_date <= ?
           ORDER BY a.warranty_date ASC""",
        (today.isoformat(), end_date)
    ).fetchall()
    db.close()
    return Response(data=[dict(r) for r in rows]).model_dump()


@router.get("/{asset_id}")
def get_asset(asset_id: int, user: dict = Depends(get_current_user),
              scope: dict = Depends(require_dept_scope())):
    db = get_db()
    row = db.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    db.close()
    if not row:
        return Response(code=1, message="资产不存在").model_dump()
    if scope:
        for k, v in scope.items():
            if row.get(k) != v:
                return Response(code=1, message="资产不存在").model_dump()
    return Response(data=_format_asset(row)).model_dump()


@router.get("/{asset_id}/timeline")
def asset_timeline(asset_id: int, user: dict = Depends(get_current_user)):
    """资产全生命周期时间线：UNION ALL stock_records + repairs + scraps"""
    db = get_db()
    rows = db.execute(
        """SELECT s.operate_date as time, s.type, s.remark as detail, u.real_name as operator
           FROM stock_records s LEFT JOIN users u ON s.operator_id = u.id
           WHERE s.asset_id = ?
        UNION ALL
           SELECT r.repair_date as time, '维修' as type, r.fault_desc as detail, u.real_name as operator
           FROM repairs r LEFT JOIN users u ON r.operator_id = u.id
           WHERE r.asset_id = ?
        UNION ALL
           SELECT sc.scrap_date as time, '报废' as type, sc.scrap_reason as detail, u.real_name as operator
           FROM scraps sc LEFT JOIN users u ON sc.operator_id = u.id
           WHERE sc.asset_id = ?
        ORDER BY time ASC""",
        (asset_id, asset_id, asset_id)
    ).fetchall()
    # Also include repair return events
    returns = db.execute(
        """SELECT r.return_date as time, '返修入库' as type, '维修完成，返修入库' as detail, u.real_name as operator
           FROM repairs r LEFT JOIN users u ON r.operator_id = u.id
           WHERE r.asset_id = ? AND r.return_confirmed = 1""",
        (asset_id,)
    ).fetchall()
    db.close()
    result = [{"time": r["time"], "type": r["type"], "detail": r["detail"], "operator": r["operator"]} for r in rows]
    for r in returns:
        if r["time"]:
            result.append({"time": r["time"], "type": r["type"], "detail": r["detail"], "operator": r["operator"]})
    result.sort(key=lambda x: x["time"] or "")
    return Response(data=result).model_dump()


@router.post("")
def create_asset(req: AssetCreate, user: dict = Depends(require_permission("asset:create"))):
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE id = ?", (req.category_id,)).fetchone()
    if not cat:
        db.close()
        return Response(code=1, message="分类不存在").model_dump()
    if cat["parent_id"] == 0:
        db.close()
        return Response(code=1, message="资产必须登记在二级类目下，请选择具体子分类").model_dump()
    asset_no = _gen_asset_no(db)
    method = req.depreciation_method or "straight"
    db.execute(
        """INSERT INTO assets (asset_no, name, category_id, brand, model, serial_no,
           purchase_price, purchase_date, dept_id, user_id, warehouse_id, location,
           supplier_id, warranty_date, remark, purchase_lifespan_years, depreciation_method)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (asset_no, req.name, req.category_id, req.brand, req.model, req.serial_no,
         req.purchase_price, req.purchase_date, req.dept_id, req.user_id,
         req.warehouse_id, req.location, req.supplier_id, req.warranty_date, req.remark,
         req.purchase_lifespan_years or 0, method)
    )
    db.commit()
    asset_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    _log(db, user["id"], f"新增资产 {asset_no} - {req.name}")
    db.close()
    return Response(data={"id": asset_id, "asset_no": asset_no}, message="资产登记成功").model_dump()


@router.put("/{asset_id}")
def update_asset(asset_id: int, req: AssetUpdate, user: dict = Depends(require_permission("asset:update"))):
    db = get_db()
    existing = db.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if not existing:
        db.close()
        return Response(code=1, message="资产不存在").model_dump()
    fields = {k: v for k, v in req.model_dump().items() if v is not None}
    if fields:
        fields["update_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sets = ", ".join(f"{k} = ?" for k in fields)
        vals = list(fields.values()) + [asset_id]
        db.execute(f"UPDATE assets SET {sets} WHERE id = ?", vals)
        db.commit()
        _log(db, user["id"], f"更新资产 {existing['asset_no']}")
    db.close()
    return Response(message="更新成功").model_dump()


@router.delete("/{asset_id}")
def delete_asset(asset_id: int, user: dict = Depends(require_permission("asset:delete"))):
    db = get_db()
    existing = db.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if not existing:
        db.close()
        return Response(code=1, message="资产不存在").model_dump()
    record_count = db.execute("SELECT COUNT(*) FROM stock_records WHERE asset_id = ?", (asset_id,)).fetchone()[0]
    repair_count = db.execute("SELECT COUNT(*) FROM repairs WHERE asset_id = ?", (asset_id,)).fetchone()[0]
    if record_count > 0 or repair_count > 0:
        db.close()
        return Response(code=1, message=f"该资产有 {record_count} 条出入库记录和 {repair_count} 条维修记录，请先清理相关记录再删除").model_dump()
    _log(db, user["id"], f"删除资产 {existing['asset_no']} - {existing['name']}")
    db.execute("DELETE FROM assets WHERE id = ?", (asset_id,))
    db.commit()
    db.close()
    return Response(message="删除成功").model_dump()


@router.post("/calculate-depreciation")
def calculate_depreciation(user: dict = Depends(require_permission("depreciation:calculate"))):
    """遍历所有 in_stock/in_use 资产，按分类配置计算折旧"""
    db = get_db()
    rows = db.execute(
        "SELECT * FROM assets WHERE status IN ('in_stock','in_use')"
    ).fetchall()
    # 加载所有分类折旧配置
    configs = {}
    config_rows = db.execute("SELECT * FROM depreciation_configs").fetchall()
    for c in config_rows:
        configs[c["category_id"]] = c
    updated = 0
    for a in rows:
        cfg = configs.get(a["category_id"])
        method = cfg["method"] if cfg else (
            a["depreciation_method"] if "depreciation_method" in a.keys() and a["depreciation_method"] else "straight"
        )
        lifespan = cfg["useful_life_years"] if cfg and cfg["useful_life_years"] else (
            a["purchase_lifespan_years"] if "purchase_lifespan_years" in a.keys() and a["purchase_lifespan_years"] else 0
        )
        salvage_rate = float(cfg["salvage_rate"] or 0) if cfg else 0
        price = float(a["purchase_price"] or 0)
        if method == "once":
            monthly = float(price)
            accumulated = float(price)
            net = 0
        elif lifespan > 0:
            depreciable = price * (1 - salvage_rate)
            monthly = depreciable / (lifespan * 12)
            accumulated = float(a["accumulated_depreciation"] or 0) + monthly if "accumulated_depreciation" in a.keys() else monthly
            net = max(price - accumulated, 0)
        else:
            monthly = 0
            accumulated = float(a["accumulated_depreciation"] or 0) if "accumulated_depreciation" in a.keys() else 0
            net = max(price - accumulated, 0)
        db.execute(
            "UPDATE assets SET monthly_depreciation=?, accumulated_depreciation=?, net_value=? WHERE id=?",
            (round(monthly, 2), round(accumulated, 2), round(net, 2), a["id"])
        )
        updated += 1
    db.commit()
    _log(db, user["id"], f"折旧计算完成，更新 {updated} 条资产")
    db.close()
    return Response(data={"updated": updated}, message="折旧计算完成").model_dump()


@router.get("/qrcode/{asset_id}")
def get_qrcode(asset_id: int):
    db = get_db()
    row = db.execute("SELECT asset_no FROM assets WHERE id = ?", (asset_id,)).fetchone()
    db.close()
    if not row:
        return Response(code=1, message="资产不存在").model_dump()
    img = qrcode.make(row["asset_no"])
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")


@router.get("/import/template")
def download_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产导入模板"
    headers = ["资产名称*", "分类ID*", "品牌", "型号", "序列号", "采购价格", "采购日期", "仓库ID", "部门ID", "供应商ID", "存放位置", "保修到期", "备注"]
    ws.append(headers)
    ws.append(["ThinkPad X1 Carbon", "1", "Lenovo", "X1 Gen11", "SN-20250001", 8999, "2025-06-01", "", "", "", "A栋3层", "2028-06-01", "开发用机"])
    for i, w in enumerate([22, 10, 12, 14, 16, 12, 14, 10, 10, 10, 14, 14, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=asset_import_template.xlsx"})


@router.post("/import")
def import_assets(file: UploadFile = File(...), user: dict = Depends(require_permission("asset:import"))):
    db = get_db()
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    imported = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip() if row[0] else ""
        if not name:
            errors.append(f"第{i}行：资产名称为空，已跳过")
            continue
        try:
            category_id = int(row[1]) if row[1] else 0
        except (ValueError, TypeError):
            errors.append(f"第{i}行：分类ID格式错误，已跳过")
            continue
        if category_id <= 0:
            errors.append(f"第{i}行：分类ID无效，已跳过")
            continue
        asset_no = _gen_asset_no(db)
        purchase_price = float(row[5]) if row[5] else 0
        purchase_date = str(row[6]) if row[6] else None
        warehouse_id = int(row[7]) if row[7] else None
        dept_id = int(row[8]) if row[8] else None
        supplier_id = int(row[9]) if row[9] else None
        location = str(row[10]) if row[10] else None
        warranty_date = str(row[11]) if row[11] else None
        remark = str(row[12]) if len(row) > 12 and row[12] else None
        db.execute(
            """INSERT INTO assets (asset_no, name, category_id, brand, model, serial_no,
               purchase_price, purchase_date, status, warehouse_id, dept_id, supplier_id,
               location, warranty_date, remark)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (asset_no, name, category_id, str(row[2]) if row[2] else None,
             str(row[3]) if row[3] else None, str(row[4]) if row[4] else None,
             purchase_price, purchase_date, "in_stock",
             warehouse_id, dept_id, supplier_id, location, warranty_date, remark)
        )
        imported += 1
    db.commit()
    _log(db, user["id"], f"批量导入资产 {imported} 条")
    db.close()
    result = {"count": imported}
    if errors:
        result["errors"] = errors[:20]
    return Response(data=result, message=f"导入成功 {imported} 条" + (f"，{len(errors)} 条失败" if errors else "")).model_dump()


@router.get("/export/all")
def export_assets(user: dict = Depends(require_permission("asset:export")),
                  scope: dict = Depends(require_dept_scope())):
    db = get_db()
    conditions = []
    params = []
    if scope:
        for k, v in scope.items():
            conditions.append(f"a.{k} = ?")
            params.append(v)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    rows = db.execute(f"SELECT * FROM assets a {where} ORDER BY a.id DESC", params).fetchall()
    db.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产列表"
    headers = ["资产编号", "名称", "品牌", "型号", "序列号", "采购价格", "采购日期", "状态", "存放位置", "备注"]
    ws.append(headers)
    for r in rows:
        ws.append([r["asset_no"], r["name"], r["brand"], r["model"], r["serial_no"],
                   r["purchase_price"], r["purchase_date"], r["status"], r["location"], r["remark"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=assets.xlsx"})
